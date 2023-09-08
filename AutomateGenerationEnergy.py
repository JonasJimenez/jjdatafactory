from asyncio import exceptions
from codecs import ignore_errors
from distutils.log import error
import json
from multiprocessing.sharedctypes import Value
from operator import index
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os, logging
import sys
import shutil

try:
    listErrores= {}
    passErrorDirectory =False
    descError = ""  

    #Path Files JSON
    path_Setting = "SettingGenEnergy.json"
      
    #LOG COnfiguracion
    Log_Format = "%(levelname)s %(asctime)s - %(message)s"
    logger = logging.getLogger()
    handler = logging.FileHandler("Log\logGenEnergy" + datetime.now().strftime('%d%m%Y%H%M%S') + ".log")
    logger.addHandler(handler)

    #Cargamos el JSON de configuracion
    with open(path_Setting, "r") as read_file:
        settingJSON = json.load(read_file)

    #Validacion parametros JSON que son obligatorios
    sourceDirectory = settingJSON['sourceDirectory']
    LoaderAVMDirectory = settingJSON['LoaderAVMDirectory']
    historyDirectory = settingJSON['HistoryDirectory']

    if len(str(sourceDirectory).replace(' ','')) <= 0:
        listErrores['sourceDirectory'] = "Falta sourceDirectory"
    if len(str(LoaderAVMDirectory).replace(' ','')) <= 0:
        listErrores['LoaderAVMDirectory'] = "Falta LoaderAVMDirectory"
    if len(str(historyDirectory).replace(' ','')) <= 0:
        listErrores['historyDirectory'] = "Falta historyDirectory"
    if len(str(settingJSON['FoundTarget']).replace(' ','')) <= 0:
        listErrores['FoundTarget'] = "Falta Objetivo..."
        logger.error("Falta Fecha")
        print("Falta Parametro (FoundTarget)")

    #Comprobar directorios
    if not (os.path.exists(sourceDirectory)):
        listErrores['pathsourceDirectory'] = "No existe directorio"
        logger.error("No existe directorio: " + sourceDirectory)
        print("No existe directorio: " + sourceDirectory)
    if not (os.path.exists(LoaderAVMDirectory)):
        listErrores['pathLoaderAVMDirectory'] = "No existe directorio"
        logger.error("No existe directorio: " + LoaderAVMDirectory)
        print("No existe directorio: " + LoaderAVMDirectory)
    if not (os.path.exists(historyDirectory)):
        listErrores['pathhistoryDirectory'] = "No existe directorio"
        logger.error("No existe directorio: " + historyDirectory)
        print("No existe directorio: " + historyDirectory)

    #Error en los Directorios
    if len(listErrores) > 0:
        raise ValueError("Error en los directorios.")
    else:
        passErrorDirectory = True

    #Leemos Directorio sourceDirectory
    ListSourceDirectory = os.listdir(settingJSON['sourceDirectory'])

    #LOG
    logger.error("\r" + "Inicio Proceso Automatizacion " + datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
    
    print("\n" + "Inicio Proceso Automatizacion " + datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
    print("Procesando Archivos: ")

    for pathSourceCierre in ListSourceDirectory:
        if pathSourceCierre.endswith(('.xlsx','.xls')):
            #LOG
            logger.error("\n" + "Procesando Archivo: " + pathSourceCierre)
            print("Archivo: " + pathSourceCierre)

            #Path Files Excel
            path_Source = sourceDirectory + "\\" + pathSourceCierre
            logger.error("path_Source:" + path_Source)
            logger.error("FoundTarget:" + settingJSON['FoundTarget'])

            #Read File Excel para obtener la Fecha Celda
            wb = load_workbook(path_Source)
            
            countActiveSheet = 0

            #Busco las hojas visibles o activas
            for sheetVisible in wb.worksheets:
                if sheetVisible.sheet_state == "visible":
                    countActiveSheet = countActiveSheet + 1               

                if (countActiveSheet > 1):
                    raise ValueError("Archivo: " + path_Source + "\r Solo esta permitido una hoja activa/visible.")             

                #Obtengo Valor Fecha de la Celda            
                sheet = sheetVisible #wb.active

                startDatetime = pathSourceCierre.replace(".xlsx","") #sheet[settingJSON['FoundTarget']].value
                logger.error("startDatetime: " + str(startDatetime))

                #Obtengo el Formato para la Fecha
                formatDate = "%d/%m/%y"

                #Reemplazo el Patron que se va borrar para obtener la Fecha Celda
                DelPattern = str(settingJSON['DelPattern']).split('|')

                #for patternDate in DelPattern:
                    #startDatetime = str(startDatetime).replace(patternDate, '')
                    #startDatetime = startDatetime.replace(' ', '')

                print("Fecha: " + startDatetime)
                
                #Se utilizara para generar los Exceles por Mes y Dia
                nameFile = datetime.today()

                try:
                    nameFile = datetime.strptime(startDatetime+sheetVisible, formatDate)
                except Exception as e:
                    raise ValueError("Archivo: " + path_Source + "\r" + str(e))

                #Inicialize List Of DataFrames at Excel Source
                listDf = {}
                findTitle = 0
                findFlagRow = 0
                exitFind = False

                #Dimensiones de la Hoja
                row_min = sheet.min_row
                row_max = sheet.max_row
                col_min = sheet.min_column
                col_max = sheet.max_column

                #For JSON
                listJSON = settingJSON['Table']

                print("[ Iniciando Automatizacion Tablas ]")

                for tablas in listJSON:
                    parameters = tablas['Parameter']
                    nameTable = str(tablas['Name'])

                    logger.error("\n" + "nameTable: " + nameTable + "\n")
                    print(nameTable)

                    for parameter in parameters:
                        cell_range = sheet[parameter['Range']]
                        
                        addRowHeader = int(str(parameter['AddRowHeader']))
                        logger.error("row_min: " + str(row_min))
                        logger.error("row_max: " + str(row_max))

                        for row in range(row_min, row_max):
                            if exitFind == True:
                                break
                            for cell in cell_range:
                                #Busco la celda que contenga el Parametro TITLE, que es el Inicio de la Tabla
                                if (str(cell[row].value).replace(' ','').upper() == str(parameter['Title']).replace(' ','').upper()):                                
                                    findTitle = row + 1 + addRowHeader
                                    exitFind = True
                                    break 
                        
                        logger.error("Fila Title: " + str(findTitle))
                        logger.error("Title: " + str(parameter['Title']).replace(' ','').upper())

                        #Se lee EXCEL de acuerdo a la RUTA y Directorio SOURCE
                        df = pd.read_excel(path_Source, 
                                    sheet_name = sheet.title,   
                                    header = findTitle,
                                    usecols = parameter['Range'])
                                                                
                        logger.error("Archivo: " + path_Source)
                        logger.error("Fila Title: " + str(findTitle) + "  Title: " + str(parameter['Title']).replace(' ','').upper())

                        logger.error("DataFrame Columns: " + df.columns[0])

                        #Relleno las Celdas Combinadas 
                        if parameter['FillMerge'] == "Yes":
                            df = df.ffill()                                       
                            
                        #Eliminacion de Filas de acuerdo al Numero de Columna y Valor de Celda en misma posicion
                        if len(str(parameter['DelColPattern'])) > 0:
                            DelColPattern = str(parameter['DelColPattern']).split('*')

                            for pattern in DelColPattern:
                                DelColPattern2 = pattern.split('|')

                                arrayIndexDelCol = []
                                for i in range(len(df)):
                                    if str(df.iloc[i,int(DelColPattern2[0])]).replace(' ', '').upper() == str(DelColPattern2[1]).replace(' ', '').upper():
                                    arrayIndexDelCol.append(i)

                                logger.error("Numero de Columnas a Eliminar: " + str(len(arrayIndexDelCol)))
                                logger.error(arrayIndexDelCol)

                                #Checar que tenga el arreglo mayor a CERO para eliminar las FIlas de acuerdo al Numero de Columna y Valor de Celda en misma posicion
                                if (len(arrayIndexDelCol) > 0):
                                    df = df.drop(arrayIndexDelCol).reset_index(drop=True)

                                arrayIndexDelCol.clear()
                                DelColPattern2.clear()
                            
                        #Quita las Columnas que no tienen Valores, por ejemplo "Unnamed"
                        if parameter['RemoveUnnamed'] == "Yes":
                            remove_cols = [col for col in df.columns if 'Unnamed' in col]
                            df.drop(remove_cols, axis='columns', inplace=True)

                        #Validacion si el DATAFRAME es DIFERENTE de VACIO
                        if not df.empty :
                            #Variable para guardar el Index de la ultima Fila a exportar, de acuerdo a la variable MaxRow y FlagRow
                            indexEnd = 0

                            # Eliminar las Filas que la columna [0] tenga NaN
                            df = df.dropna(subset=[df.columns[0]]).reset_index(drop=True)

                            #Validamos si existe valor para FlagRow, Bandera para localizar el RowIndex de una Valor, Ultima Fila a Exportar del Dataframe
                            try:
                                if len(str(parameter['FlagRow']).replace(' ','')) > 0:
                                    #Se obtiene el indice de la Columna 0            
                                    indexEnd = (df.index[df[df.columns[0]] == parameter['FlagRow']])[0] #- 1            
                                else:
                                    if int(parameter['MaxRow']) == 0:           
                                        for i in range(len(df)):  #Basicamente lee la ultima Fila, por que no hay parametro 'FlagRow' ni 'MaxRow"
                                            indexEnd = i                                                       
                                    else: #Asigna indexEnd con el parametro del JSON 'MaxRow'
                                        indexEnd = int(parameter['MaxRow'])
                            except Exception as e:
                                logger.error("Error Valor Parametro: FlagRow " +  str(parameter['FlagRow']) + ", " + str(e))
                                raise ValueError("Archivo: " + path_Source + "\r Tabla: " + nameTable +  "\r Error Valor Parametro: FlagRow " + str(parameter['FlagRow']) + ", "  + str(e))

                            arrayHeader = []
                            countHeader = 64 #ASCILL ABC

                            #Asignamos Encabezado aquellas Tablas que no tienen en Excel, Encabezado de acuerdo al ABC
                            if parameter['AddHeader'] == "Yes":
                                for header in df.columns:
                                    countHeader = countHeader + 1
                                    arrayHeader.append(chr(countHeader))

                                df.columns = arrayHeader

                            # Indico el rango de Filas que se necesita, solo esportamos del header hasta la ultima fila de acuerdo a los parametros MaxRow y FlagRow
                            df = df.iloc[0:indexEnd]                  

                            #Eliminar duplicados
                            df = df.drop_duplicates()

                            #Formato a la Fecha Obtenida del Excel
                            dateColumn = datetime.strptime(startDatetime, formatDate)

                            #Adding Column Datetime
                            df.insert(0, 'FECHA', '{}/{}/{}'.format(dateColumn.month,dateColumn.day,dateColumn.year), True)

                            #Eliminacion de Columnas
                            if len(str(parameter['DeleteColum'])) > 0:
                                DeleteColum = str(parameter['DeleteColum']).split('|')
                                arrayIndexDeleteColumnx = []
                                for patternDelColumn in DeleteColum:
                                    arrayIndexDeleteColumnx.append(int(patternDelColumn))

                                df.drop(df.columns[arrayIndexDeleteColumnx], axis = 'columns', inplace = True)

                            if len(str(parameter['ArrayColumn'])) > 0:
                                arrayLevel = []
                                arrayFluid = []
                                for i in range(len(df)):    
                                    arrayLevel.append(str(df.iloc[i,int(str(parameter['ArrayColumn']))]).split('\'')[0])
                                    arrayFluid.append(str(df.iloc[i,int(str(parameter['ArrayColumn']))]).split('\'')[1].split('"')[0])

                                df.insert(len(df.columns), 'LEVEL', arrayLevel, True)
                                df.insert(len(df.columns), 'FLUID_IN', arrayFluid, True)

                            if len(str(parameter['ArrayHour'])) > 0:
                                arrayHour = []
                                for i in range(len(df)):         
                                    arrayHour.append(str(df.iloc[i,int(str(parameter['ArrayHour']))]).split(':')[0])

                                df.insert(len(df.columns), 'HOUR', arrayHour, True)

                            #Adding Dataframe in ListDF
                            listDf[nameTable] = df
                            logger.error("Registros Exportados: " + str(len(df.index)))
                            print("Registros Exportados: " + str(len(df.index)))                     

                        else:
                            print("Registros Exportados: 0") 

                            #Adding Dataframe in ListDF
                            listDf[nameTable] = df
                            logger.error("Dataframe VACIO: " + nameTable)

                    #Resetear Variables
                    findTitle = 0
                    exitFind  = False

        writer = pd.ExcelWriter(LoaderAVMDirectory + '\\' + 'ReporteCierreDiario_'+ '{}_{}_{}'.format(dateColumn.day,dateColumn.month,dateColumn.year) + '.xlsx')

        for key, result in listDf.items():        
                result.to_excel(writer, sheet_name = key, index= False)

        writer.save()   

        try:
            #Mover los archivos al folder History, si el archivo ya existe lo sobreescribe
            if len(str(historyDirectory).replace(' ','')) > 0:
                path_Source  = sourceDirectory + "\\" + pathSourceCierre
                path_History = historyDirectory + "\\" + pathSourceCierre
                shutil.move(path_Source, path_History)
                logger.error("Se Movio el Archivo: " + path_Source)
                logger.error("A Carpeta History: " + path_History)
        except exceptions as e:
            raise ValueError(format(e))
except ValueError as e:
    descError = str(e)   
    print("Error: {}".format(e))
    logger.error("Error: {}".format(e))        
except exceptions as e:
    descError = str(e)
    print("Error: {}".format(e))
    logger.error("Error: {}".format(e))
finally:
    if (len(descError) > 0):
            sys.exit("Error: {}".format(descError))     
    elif (passErrorDirectory):
        #Validar con el usuario si es conveniente generar error cuando no hay archivos para procesar o finalizar con Exito
        if  (len(ListSourceDirectory) <= 0):
            logger.error("No hay archivos que procesar.")
            logger.error("Ruta: " + sourceDirectory)
            print("No hay archivos que procesar.")
            print("Ruta: " + sourceDirectory)
            sys.exit("No hay archivos que procesar.")
        else:
            logger.error("")
            logger.error("Fin Proceso Automatizacion " + datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
            print("Fin Proceso Automatizacion " + datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
            sys.exit() 

            '''
            try:
                #Mover los archivos al folder History, si el archivo ya existe lo sobreescribe
                if len(str(historyDirectory).replace(' ','')) > 0:
                    for pathSourceCierre in ListSourceDirectory:
                        if pathSourceCierre.endswith(('.xlsx','.xls')):
                            #Path Files Excel
                            path_Source  = sourceDirectory + "\\" + pathSourceCierre
                            path_History = historyDirectory + "\\" + pathSourceCierre
                            shutil.move(path_Source, path_History)
                            logger.error("Se Movio el Archivo: " + path_Source)
                            logger.error("A Carpeta History: " + path_History)
            except exceptions as e:
                logger.error("Error: {}".format(e))
                sys.exit("Error: {}".format(e))
            finally:
                logger.error("")
                logger.error("Fin Proceso Automatizacion " + datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
                print("Fin Proceso Automatizacion " + datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
                sys.exit()      
            '''